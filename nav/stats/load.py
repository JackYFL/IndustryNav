"""Adapters from raw data sources into a normalized per-run row format.

Three input sources funnel into the same row schema so the bootstrap /
permutation / Spearman primitives downstream can stay source-agnostic:

- :func:`discover_grid_runs` — walks ``outputs/<scene>/<point>/<model>[
  _novision]/seed<k>/results.csv`` produced by ``run_benchmark_grid.py``.
- :func:`xlsx_to_per_run_rows` — parses the archived
  ``IndustryNav_extra_agents_results.xlsx`` single-run sweep into the
  same row dict shape (was a standalone ``xlsx_to_per_run.py``).
- :func:`load_per_run_csv` — reads back a previously-emitted per_run.csv,
  so downstream stages can be re-run without redoing the walk.

The legacy experiment_results_v6.csv path uses a different (model-
keyed) shape — see :func:`load_experiment_v6_by_model` — because that
source doesn't carry the (scene, point) identity the canonical schema
requires.
"""

from __future__ import annotations

import csv
import math
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import numpy as np

from nav.config import EVAL_COLLISION_PX_THRESH
from nav.eval.collision import compute_collision_rate
from nav.eval.metrics import compute_success_efficiency_distance


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_float_or_nan(v) -> float:
    """Coerce ``v`` to ``float``; return NaN on any parse / non-finite failure."""
    try:
        x = float(v)
        return x if math.isfinite(x) else float("nan")
    except (TypeError, ValueError):
        return float("nan")


def _normalize_ratio(v) -> Optional[float]:
    """Normalize a metric to a 0-1 ratio.

    The xlsx mixes percentages (0-100) and ratios (0-1) inconsistently
    within and across sheets; anything > 1.0001 is treated as a percentage
    and divided by 100. Returns None for missing / non-numeric / negative.
    """
    if v is None:
        return None
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x):
        return None
    if x > 1.0001:
        x = x / 100.0
    if x < 0:
        return None
    return x


def _vision_input_bool(value) -> bool:
    """Coerce a vision_input column value (str, bool, int) to a bool."""
    return str(value).lower() in {"1", "true", "yes", "on"}


# ---------------------------------------------------------------------------
# Grid output tree → per-run rows
# ---------------------------------------------------------------------------

def discover_grid_runs(
    outputs_root: Path,
    *,
    models_filter: Optional[set] = None,
    vision_filter: Optional[set] = None,
) -> List[dict]:
    """Walk the grid output tree and return one row per (scene, point, model, seed) cell.

    Layout expected::

        outputs/<scene>/<point>/<model_short>[_novision]/seed<k>/{results,agent_actions}.csv

    For each cell we re-derive SR, distance-ratio, and collision-rate from
    the per-frame ``agent_actions.csv`` (matching what
    ``eval_metrics.evaluate_run`` would produce) so the rebuttal numbers
    stay in lockstep with the paper's evaluation pipeline. Warning-rate
    cannot be reconstructed (grid runs save depth PNG visualizations, not
    the raw NPYs), so it lands as NaN — the xlsx path supplies it when
    available.

    ``vision_filter`` is the set of strings ``{"on"}``, ``{"off"}``, or
    both; ``None`` means accept everything.
    """
    rows: List[dict] = []
    if not outputs_root.exists():
        return rows
    if vision_filter is None:
        vision_filter = {"on", "off"}

    for scene_dir in outputs_root.iterdir():
        if not scene_dir.is_dir() or scene_dir.name.startswith("_"):
            continue
        scene_name = scene_dir.name
        for point_dir in scene_dir.iterdir():
            if not point_dir.is_dir():
                continue
            point_id = point_dir.name
            for model_dir in point_dir.iterdir():
                if not model_dir.is_dir():
                    continue
                model_short = model_dir.name
                if model_short.endswith("_novision"):
                    vision = False
                    model_short_clean = model_short[: -len("_novision")]
                else:
                    vision = True
                    model_short_clean = model_short
                if (vision and "on" not in vision_filter) or (
                    not vision and "off" not in vision_filter
                ):
                    continue

                for seed_dir in model_dir.iterdir():
                    if not seed_dir.is_dir() or not seed_dir.name.startswith("seed"):
                        continue
                    rcsv = seed_dir / "results.csv"
                    if not rcsv.exists():
                        continue
                    rows.extend(
                        _rows_from_seed_dir(
                            seed_dir, rcsv,
                            scene_name=scene_name,
                            point_id=point_id,
                            model_short_clean=model_short_clean,
                            vision=vision,
                            models_filter=models_filter,
                        )
                    )
    return rows


def _rows_from_seed_dir(
    seed_dir: Path,
    rcsv: Path,
    *,
    scene_name: str,
    point_id: str,
    model_short_clean: str,
    vision: bool,
    models_filter: Optional[set],
) -> List[dict]:
    """Yield row dicts for the contents of one ``seed<k>/`` directory."""
    actions_csv = seed_dir / "agent_actions.csv"
    if actions_csv.exists():
        try:
            sr_canon, eff_steps, dr = compute_success_efficiency_distance(actions_csv)
            _, _, cr = compute_collision_rate(actions_csv, EVAL_COLLISION_PX_THRESH)
        except Exception:
            sr_canon, eff_steps, dr, cr = 0, 0, float("nan"), float("nan")
    else:
        sr_canon, eff_steps, dr, cr = 0, 0, float("nan"), float("nan")

    out: List[dict] = []
    with open(rcsv) as f:
        for r in csv.DictReader(f):
            model = r.get("model") or ""
            if models_filter and model not in models_filter:
                continue
            dist = _to_float_or_nan(r.get("distance_px"))
            if not math.isfinite(dist):
                continue
            out.append({
                "scene_name": r.get("scene_name") or scene_name,
                "point_id": r.get("point_id") or point_id,
                "model": model,
                "model_short": model_short_clean,
                "vision_input": vision,
                "seed_id": r.get("seed_id") or seed_dir.name.replace("seed", ""),
                # Canonical SR comes from the world-distance evaluation helper.
                "success": int(sr_canon),
                "distance_px": dist,
                "distance_ratio": dr,
                "collision_rate": cr,
                "warning_rate": float("nan"),  # grid runs don't preserve raw depth
                "efficiency_steps": eff_steps,
                "steps_taken": int(_to_float_or_nan(r.get("steps_taken")) or 0),
            })
    return out


# ---------------------------------------------------------------------------
# xlsx single-run sweep → per-run rows
# ---------------------------------------------------------------------------

def xlsx_to_per_run_rows(xlsx_path: Path) -> List[dict]:
    """Parse the archived xlsx sweep into per-run rows in the canonical schema.

    Sheet layout: one sheet per scene (12 total). Each sheet has a header
    row (``Model | Success ratio | Distance Ratio | Efficiency | Collision
    Ratio | Warning Ratio``) and data rows where the first column carries
    a "Point N" label on the row that starts each (scene, point) section.

    The xlsx is single-condition vision-on; ``vision_input=True`` and
    ``seed_id="0"`` are tagged onto every row to match the canonical schema.
    """
    from openpyxl import load_workbook  # lazy import — openpyxl is heavy

    wb = load_workbook(xlsx_path, data_only=True)
    all_rows: List[dict] = []
    for ws in wb.worksheets:
        scene = ws.title.strip().lower()
        all_rows.extend(_parse_xlsx_sheet(ws, scene))
    return all_rows


def _parse_xlsx_sheet(ws, scene_name: str) -> List[dict]:
    """Parse one xlsx worksheet into per-run rows."""
    # Locate header row by scanning for a 'Model' cell.
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue
        if any(isinstance(c, str) and c.strip().lower() == "model" for c in row):
            header_row_idx = i
            break
    if header_row_idx is None:
        return []

    rows: List[dict] = []
    current_point: Optional[str] = None
    for row in list(ws.iter_rows(values_only=True))[header_row_idx + 1:]:
        cells = list(row) + [None] * (7 - len(row)) if row else [None] * 7
        point_label, model, sr, dr, eff, cr, wr = cells[:7]

        # Section label lives on the first row of each (scene, point) section only.
        if isinstance(point_label, str) and point_label.strip().lower().startswith("point"):
            current_point = point_label.strip().lower().replace(" ", "")

        if not isinstance(model, str):
            continue
        m = model.strip()
        if not m or "human" in m.lower():  # drop the oracle baseline
            continue
        if sr is None and dr is None and cr is None and wr is None:
            continue
        if current_point is None:
            continue  # defensive — should not happen for well-formed sheets

        try:
            success = int(round(float(sr))) if sr is not None else None
        except (TypeError, ValueError):
            success = None

        try:
            eff_int = int(round(float(eff))) if eff is not None else 0
        except (TypeError, ValueError):
            eff_int = 0

        rows.append({
            "scene_name": scene_name,
            "point_id": current_point,
            "model": m,
            "model_short": m.split("/")[-1],
            "vision_input": True,        # xlsx is single-condition vision-on
            "seed_id": "0",              # one run per cell, no replicates
            "success": success,
            "distance_px": float("nan"),  # not recorded in the xlsx
            "distance_ratio": _normalize_ratio(dr),
            "collision_rate": _normalize_ratio(cr),
            "warning_rate": _normalize_ratio(wr),
            "efficiency_steps": eff_int,
            "steps_taken": eff_int,
        })
    return rows


# ---------------------------------------------------------------------------
# per_run.csv → per-run rows
# ---------------------------------------------------------------------------

def load_per_run_csv(csv_path: Path) -> List[dict]:
    """Load a previously-emitted per_run.csv into the canonical row schema.

    Used when downstream stages need to re-process rows without redoing the
    grid walk / xlsx parse. ``success`` may be float (from merged
    seed-averaged rows in [0,1]) or int 0/1 (single-run cells); the
    downstream bootstrap / permutation treat it as a real-valued metric
    either way.
    """
    if not csv_path.exists():
        raise FileNotFoundError(f"per_run.csv not found: {csv_path}")

    rows: List[dict] = []
    with open(csv_path) as f:
        for r in csv.DictReader(f):
            try:
                success = float(r["success"])
            except (TypeError, ValueError):
                continue
            if not math.isfinite(success):
                continue
            rows.append({
                "scene_name": r["scene_name"],
                "point_id": r["point_id"],
                "model": r["model"],
                "model_short": r.get("model_short") or r["model"].split("/")[-1],
                "vision_input": _vision_input_bool(r.get("vision_input", "True")),
                "seed_id": r.get("seed_id") or "0",
                "success": success,
                "distance_px": _to_float_or_nan(r.get("distance_px")),
                "distance_ratio": _to_float_or_nan(r.get("distance_ratio")),
                "collision_rate": _to_float_or_nan(r.get("collision_rate")),
                "warning_rate": _to_float_or_nan(r.get("warning_rate")),
                "efficiency_steps": _int_or_zero(r.get("efficiency_steps")),
                "steps_taken": _int_or_zero(r.get("steps_taken")),
            })
    return rows


def _int_or_zero(v) -> int:
    x = _to_float_or_nan(v)
    return int(round(x)) if math.isfinite(x) else 0


# ---------------------------------------------------------------------------
# experiment_results_v6.csv → by-model arrays (partial-stats source)
# ---------------------------------------------------------------------------

def load_experiment_v6_by_model(
    csv_path: Path, success_stop_reason: str = "reached_vicinity"
) -> Dict[str, Dict[str, np.ndarray]]:
    """Load the legacy archive CSV grouped by model.

    Unlike the other loaders this returns ``{model: {"distance": np.array,
    "success": np.array, "steps": np.array}}`` — the source CSV doesn't
    carry scene/point identifiers, so the canonical (scene, point, seed)
    row schema isn't reconstructible. Used by the partial-stats pipeline.
    """
    by_model: Dict[str, Dict[str, List]] = {}
    with open(csv_path) as f:
        for row in csv.DictReader(f):
            dist = _to_float_or_nan(row.get("distance_px"))
            model = row.get("model") or ""
            if not model or not math.isfinite(dist):
                continue
            success = 1 if row.get("stop_reason", "") == success_stop_reason else 0
            d = by_model.setdefault(model, {"distance": [], "success": [], "steps": []})
            d["distance"].append(dist)
            d["success"].append(success)
            try:
                d["steps"].append(int(row.get("steps_taken") or "0"))
            except ValueError:
                d["steps"].append(0)
    return {m: {k: np.asarray(v) for k, v in d.items()} for m, d in by_model.items()}


# ---------------------------------------------------------------------------
# Persistence
# ---------------------------------------------------------------------------

def write_per_run_csv(rows: List[dict], out_path: Path) -> None:
    """Write rows back to a per_run.csv with column order from ``rows[0]``."""
    if not rows:
        raise ValueError("no rows to write")
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def summarize_coverage(rows: Iterable[dict]) -> str:
    """Return a multi-line text summary of model + scene coverage."""
    rows = list(rows)
    by_model = Counter(r["model"] for r in rows)
    by_scene = Counter(r["scene_name"] for r in rows)
    parts = [
        f"{len(rows)} rows, {len(by_model)} models, {len(by_scene)} scenes",
        "Models:",
    ]
    for m, n in by_model.most_common():
        parts.append(f"  {m:<45s} N={n}")
    parts.append("Scenes:")
    for s, n in sorted(by_scene.items()):
        parts.append(f"  {s:<15s} N={n}")
    return "\n".join(parts)
