"""Batch aggregation across many runs.

:func:`aggregate_runs` consumes an iterable of run directories, calls
:func:`nav.eval.metrics.evaluate_run` on each, and returns one row per
input. Failures become rows with the ``error`` field populated rather
than raising — the caller decides how to surface them.

:func:`write_aggregate_xlsx` is the xlsx-emission half. It depends on
``openpyxl`` being installed; if not, it falls back to CSV. Used by
``nav.scripts.aggregate_eval``.
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from pathlib import Path
from typing import Iterable, List, Optional

import numpy as np

from nav.eval.metrics import EvaluateOptions, evaluate_run


#: Schema for the per-run rows produced by :func:`aggregate_runs`.
AGGREGATE_ROW_FIELDS: List[str] = [
    "scene_name", "point_id", "model",
    "success_ratio", "efficiency_steps", "distance_ratio",
    "final_distance_world", "final_distance_px", "stop_reason",
    "warning_rate", "warning_steps", "total_steps",
    "collision_rate", "collision_steps", "forward_steps",
    "error", "input_dir",
]


def _row_axes_from_path(input_dir: Path) -> dict:
    """Infer (scene_name, point_id, model) from a path ending in those parts."""
    parts = input_dir.parts
    return {
        "scene_name": parts[-3] if len(parts) >= 3 else "",
        "point_id": parts[-2] if len(parts) >= 2 else "",
        "model": parts[-1] if len(parts) >= 1 else "",
    }


def _error_row(input_dir: Path, error: Exception) -> dict:
    base = _row_axes_from_path(input_dir)
    base.update({
        "input_dir": str(input_dir),
        "success_ratio": 0,
        "efficiency_steps": 0,
        "distance_ratio": 0.0,
        "final_distance_world": None,
        "final_distance_px": None,
        "stop_reason": "",
        "warning_rate": 0.0,
        "warning_steps": 0,
        "total_steps": 0,
        "collision_rate": 0.0,
        "collision_steps": 0,
        "forward_steps": 0,
        "error": str(error),
    })
    return base


def aggregate_runs(
    input_dirs: Iterable[Path],
    opts: Optional[EvaluateOptions] = None,
) -> List[dict]:
    """Evaluate each directory and return one row per input.

    Failures are converted to rows with the ``error`` field set, so the
    output length always matches the input length and the caller can
    cleanly separate ok vs errored rows by ``row.get("error")``.
    """
    opts = opts or EvaluateOptions()
    rows: List[dict] = []
    for input_dir in input_dirs:
        input_dir = Path(input_dir)
        if not input_dir.is_dir():
            continue
        try:
            metrics = evaluate_run(input_dir, opts)
            row = {**_row_axes_from_path(input_dir), **metrics, "error": ""}
        except Exception as exc:  # noqa: BLE001 — surface any read failure
            row = _error_row(input_dir, exc)
        rows.append(row)
    return rows


def write_aggregate_csv(out_path: Path, rows: List[dict]) -> None:
    """Write per-run rows to a flat CSV using :data:`AGGREGATE_ROW_FIELDS`."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=AGGREGATE_ROW_FIELDS, extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(rows)


def summarize_by(rows: List[dict], axis: str) -> List[dict]:
    """Group rows by ``axis`` and return mean metrics per group.

    Groups are ordered by ``axis`` key. ``axis`` is typically
    ``"scene_name"`` for per-scene rollups; pass any column name that
    discriminates the grouping you want.
    """
    groups: dict = defaultdict(list)
    for row in rows:
        if row.get("error"):
            continue
        groups[row.get(axis, "")].append(row)

    summary_rows = []
    for key in sorted(groups):
        group = groups[key]
        n = len(group)
        success = sum(int(r["success_ratio"]) for r in group)
        dist_vals = [r["distance_ratio"] for r in group if r["distance_ratio"] not in (None, "")]
        warn_vals = [float(r["warning_rate"]) for r in group if r["warning_rate"] not in (None, "")]
        coll_vals = [float(r["collision_rate"]) for r in group]
        eff_vals = [float(r["efficiency_steps"]) for r in group]
        summary_rows.append({
            axis: key,
            "N": n,
            "success_rate_%": round(success / n * 100, 2) if n else 0.0,
            "distance_ratio_%": (
                round(float(np.nanmean(dist_vals)) * 100, 2) if dist_vals else float("nan")
            ),
            "efficiency_steps": round(float(np.mean(eff_vals)), 2) if eff_vals else 0.0,
            "warning_rate_%": (
                round(float(np.mean(warn_vals)) * 100, 2) if warn_vals else float("nan")
            ),
            "collision_rate_%": (
                round(float(np.mean(coll_vals)) * 100, 2) if coll_vals else 0.0
            ),
        })
    return summary_rows


def write_aggregate_xlsx(
    out_path: Path,
    rows: List[dict],
    summary_axis: str = "scene_name",
) -> bool:
    """Write per-run rows + a summary sheet to ``out_path``.

    Returns True on xlsx success, False if openpyxl is unavailable (caller
    can then fall back to :func:`write_aggregate_csv`).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Detailed Results"
    ws1.append(AGGREGATE_ROW_FIELDS)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for row in rows:
        ws1.append([row.get(h, "") for h in AGGREGATE_ROW_FIELDS])
    for col in ws1.columns:
        ws1.column_dimensions[get_column_letter(col[0].column)].width = 18

    summary = summarize_by(rows, axis=summary_axis)
    if summary:
        ws2 = wb.create_sheet(f"Summary by {summary_axis}")
        s_headers = list(summary[0].keys())
        ws2.append(s_headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="70AD47")
        for srow in summary:
            ws2.append([srow.get(h, "") for h in s_headers])
        for col in ws2.columns:
            ws2.column_dimensions[get_column_letter(col[0].column)].width = 20

    wb.save(out_path)
    return True
